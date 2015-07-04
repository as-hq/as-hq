import numpy as np
import random
from AS.iterable import ASIterable
import matplotlib.pyplot as plt
from openpyxl import load_workbook
import string
from clusters import *

def arr(lst):
	return ASIterable(lst)

def space(lst, sp):
	lst2 = map((lambda x: prefixPush(x, ["" for _ in range(sp)])), lst)
	return flat(lst2)[:-sp]

def flat(lst):
	return [item for sublist in lst for item in sublist]

def prefixPush(elem, lst):
	lst.insert(0, elem) 
	return lst

def every(lst, k):
	return lst[0::k]

def sumWay(lst, axis):
	return np.sum(lst, axis).tolist()

def multiply(lst1, lst2):
	return (np.multiply(lst1, lst2)).tolist()

def reshape(lst,axis1,axis2):
	return np.array(lst).reshape((axis1, axis2)).tolist() 

def transpose(lst):
	return np.array(lst).transpose().tolist()

def sumSquares(lst):
	return sum(np.array(lst)**2)

def rand(m=1,n=1,upperbound=1):
	if n==1 and m != 1:
		return ASIterable(np.random.rand(m)*randint(1,upperbound))
	elif m==1 and n!=1:
		return ASIterable(np.random.rand(m)*randint(1,upperbound)).transpose()
	elif m==1 and n==1:
		return np.random_sample*random.randint(1,upperbound)
	else: return ASIterable(np.random.rand(m,n)*random.randint(1,upperbound))


def readSheet(filePath, sheetName=None):
    wb = load_workbook(filePath, read_only=True)
    wbData = load_workbook(filePath, read_only=True, data_only=True)
    wa = wb.active
    wa2 = wbData.active
    if (sheetName!=None):
        wa=wb[sheetName]
        wa2=wbData[sheetName]
    locs = []
    exprs = []
    vals = []
    for row in wa.rows:
        for cell in row:
            if cell.row!=None and cell.column!=None:
                index = cellToIndex(cell)
                expr = str(exprToPython(cell))
                if expr!="":
                    locs.append(index)
                    exprs.append(expr)
    for row in wa2.rows:
        for cell in row:
            if cell.row!=None and cell.column!=None:
              vals.append(exprToPython(cell))
    return {"excelLocs": locs, "excelExprs": exprs, "excelVals": vals}


def excelColToNum(col):
    num = 0
    for c in col:
        if c in string.ascii_letters:
            num = num * 26 + (ord(c.upper()) - ord('A')) + 1
    return num

def cellToIndex(cell):
    return [excelColToNum(cell.column), cell.row]

def exprToPython(cell):
    # eventually need to map excel functions to python functions
    if cell.value == None or cell.value=="":
            return ""
    try:
        s=cell.value.encode('ascii', 'ignore')
        if (s[0]!='='):
            return s
        else:
            return s[1:] 
    except:
        return cell.value
    

