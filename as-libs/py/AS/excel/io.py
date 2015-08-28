from AS.excel.pycel.excelcompiler import *
from AS.excel.pycel.excellib import * # mapping from excel to python
from openpyxl import load_workbook

def readSheet(filePath, sheetName=None):
	wb = load_workbook(filePath, read_only=True)
	wbData = load_workbook(filePath, read_only=True, data_only=True)
	wa = wb.active
	wa2 = wbData.active
	if (sheetName!=None):
		wa=wb[sheetName]
		wa2=wbData[sheetName]
	locs = []
	exprs = []
	vals = []
	for row in wa.rows:
		for cell in row:
			if cell.row!=None and cell.column!=None:
				index = cellToIndex(cell)
				expr = str(exprToPython(cell))
				if expr!="":
					locs.append(index)
					exprs.append(expr)
	for row in wa2.rows:
		for cell in row:
			if cell.row!=None and cell.column!=None:
			  vals.append(exprToPython(cell))
	return {"excelLocs": locs, "excelExprs": exprs, "excelVals": vals}

def excelColToNum(col):
	num = 0
	for c in col:
		if c in string.ascii_letters:
			num = num * 26 + (ord(c.upper()) - ord('A')) + 1
	return num

def cellToIndex(cell):
	return [excelColToNum(cell.column), cell.row]

def exprToPython(cell):
	# eventually need to map excel functions to python functions
	if cell.value == None or cell.value=="":
			return ""
	try:
		s=cell.value.encode('ascii', 'ignore')
		if (s[0]!='='):
			return s
		else:
			return s[1:] 
	except:
		return cell.value
	
def evalExcel(xp):
    e = shunting_yard(xp)
    G,root = build_ast(e)
    newXp = root.emit(G,context=None)
    vols = [node.isVolatile for node in G.nodes()]
    hasVolatile = False
    for v in vols:
        if v:
            hasVolatile = True
    return [newXp,hasVolatile]
